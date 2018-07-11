import Members
import xlsxwriter 
import datetime
import string
from openpyxl import load_workbook

def oneMember( sheet, rowIndex):
    member = Members.Members()
    info = []
    for row_cells in sheet.iter_rows(min_row=rowIndex, max_row=rowIndex):
        for cell in row_cells:
            info.append(cell.value)
    member.count = info[0]
    member.department = info[1]
    member.name = info[2]
    member.lastDate = info[3]
    return member
def members(sheet):
    members = []
    rowIndex = 2
    while rowIndex <= sheet.max_row:
        member = oneMember(sheet, rowIndex)
        members.append(member)
        rowIndex = rowIndex + 1
    return members

def depMembers(members, depName):
    depMembers = []
    for member in members:
        if member.department == depName:
            depMembers.append(member)
    return depMembers

def getGap(lastWeekiesDate, today):
   delta = today - lastWeekiesDate
   return delta.days

def getNoPto(members):
    numOoo = int(input("How many will not be in office next week?")) 
    count = 0
    while count < numOoo:
        name = input("Name: ")
        for member in members:
            if member.name.casefold() == name.casefold():
                members.remove(member)      
        count = count + 1
    return members
def getCandidates(members):
    candidates = getNoPto(members)
    #remove who was a weekie recently
    candidatesCopy =candidates.copy()
    for candidate in candidatesCopy:
        if getGap(candidate.lastDate, datetime.datetime.today()) < 35:
            candidates.remove(candidate)
    return candidates

def updateInfoInFile(weekieOne, weekieTwo, sheet):
    members = []
    rowIndex = 2
    dateIndex = 3
    countIndex = 0
    nameIndex = 2
    for row_cells in sheet.iter_rows(min_row=rowIndex, max_row=sheet.max_row):
        if row_cells[nameIndex].value == weekieOne.name or row_cells[nameIndex].value == weekieTwo.name:
            
            row_cells[dateIndex].value = (datetime.datetime.today()+datetime.timedelta(days = 8)).strftime('%Y-%m-%d')
            row_cells[countIndex].value = row_cells[countIndex].value + 1
      